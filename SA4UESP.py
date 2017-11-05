import SAdef
import random
import time
import math
import copy
import openpyxl

t = 21 #the number of timeslots
stu_file = 'E:\literature\done\SA4UESP\dataset\yor-f-83.stu' 
crs_file = 'E:\literature\done\SA4UESP\dataset\yor-f-83.crs'
dataset = 'yor83'

wb = openpyxl.load_workbook('E:\literature\done\SA4UESP\SA4UESP_RESULTS.xlsx')
ws = wb['Sheet1']
arow = ws.max_row + 1
ws.cell(row = arow, column = 1).value = dataset
print ws.cell(row = arow, column = 1).value
que_s = []
que_iteration = []
que_timecost = []

'''create Cij'''
cij = SAdef.create_cij(stu_file,crs_file)

n = SAdef.n #get the number of exams
m = SAdef.m #get the number of students

for z in range(10):

    '''compute an initial feasible exam timetable s_initial'''
    infeasible = True
    while infeasible: #if the s_initial timetable is infeasible, need to repeat initialization
        infeasible = False
        exam_list = [-1 for i in range(n)]
        s_initial = [[] for i in range(t)]
            #random distribute the first exam
        tt = random.randint(0,t-1)
        exam_list[0] = tt
        s_initial[tt].append(0)
            #distribute the rest of exams
        for i in range(n-1):
            p = SAdef.mark_hardest(exam_list,s_initial,cij) #choose the most conflicting one
            if p==-1: #p==-1 means there has no available timeslot for someone exam, infeasible
                infeasible = True
                break
                #check if s_initial[tt] is available for exam_list[p]
            conflict = True
            while conflict:
                tt = random.randint(0,t-1)
                if s_initial[tt]!=[]:
                    conflict = False
                    for j in s_initial[tt]:
                        if cij[p][j]!=0:
                            conflict = True
                            break
                else:
                    conflict = False
                #put exam_list[p] in s_initial[tt]
            exam_list[p] = tt
            s_initial[tt].append(p)

    '''using simulating annealing to find the best solution'''

    s_accepted = s_initial
    s_best = s_initial
    tk = 1.0
    tf = 0.1
    begin = time.time()
    elapsed_time = 0
    k = 1.02*100
    while(tk>=tf and elapsed_time < 7200):
        for it in range(int(k)):
            rand = random.random()
            if rand <= 0.2:
                s,d = SAdef.n1(s_accepted,cij)
            elif rand <= 0.6:
                s,d = SAdef.n2(s_accepted,cij)
            else:
                s,d = SAdef.n3(s_accepted,cij)
            if d < 0:
                s_accepted = s
            elif random.random() < math.exp(-d/tk):
                s_accepted = s
            if SAdef.evaluate(s_best,cij) > SAdef.evaluate(s_accepted,cij):
                s_best = s_accepted
        if int(k) == 9506:
            que_iteration.append(it)
        k = k*1.02
        tk = 0.99*tk
        elapsed_time = time.time() - begin
    que_timecost.append(elapsed_time)

    #ouput s_best
    best_s = SAdef.evaluate(s_best,cij)
    ws.cell(row = arow,column = z+2).value = best_s
    que_s.append(best_s)
ws.cell(row = arow, column = 12).value = max(que_s)
ws.cell(row = arow, column = 16).value = min(que_s)
ws.cell(row = arow, column = 13).value = sum(que_s)/len(que_s)
ws.cell(row = arow, column = 14).value = sum(que_iteration)/len(que_iteration)
ws.cell(row = arow, column = 15).value = sum(que_timecost)/len(que_timecost)
wb.save('E:\literature\done\SA4UESP\SA4UESP_RESULTS.xlsx')

